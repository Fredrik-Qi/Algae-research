# code_#1 image segmentation

'''从文件夹中读取所有图片，并从每张图片的一系列坐标开始，以100*100/33*33的大小的分割框对图像进行分割;每个点之间相距30个像素点
，每次向右移动60/30像素，将图片分割成小图片。最终以原图片名+行名(A-H)+列名(1-12)的形式保存在一个以原图片名为名的文件夹中'''

import cv2
import os
file_pathname=""  #文件路径

def read_path(file_pathname):
    for root, dirs, files in os.walk(file_pathname):
        #os.walk 
        for file in files:  
            ##print(file_pathname+'/'+file)  #test code
            if file.endswith('.png'):  #selcet picture
                #Load pictures
                img = cv2.imread(file_pathname+'/'+file)
                file_name, file_extend = os.path.splitext(file)  
                # Define the coordinates坐标 and size of the crop box

                coords = [[275,225], [275,255], [275, 285], [275,315], [275, 345], [275,375], [275, 405], 
                          [275,435], [275, 465], [275,495], [275, 525], [275,555], [275, 585], [275,615],[275, 645],[275,675]]
                width, height = 33, 33
                move_x = 30
                #当分割框包含1个衣藻时选择该组参数


                '''
                coords = [[275,225], [275, 285],[275, 345], [275, 405], [275, 465], [275, 525], [275, 585], [275, 645]]
                width, height = 100, 100
                move_x = 60
                #当分割框包含9个衣藻时选择该组参数
                '''


                # Loop through all the coordinates and crop the image
                for i, coord in enumerate(coords):
                    x, y = coord
                    for j in range(1,25): 
                        crop_img = img[y:y+height, x:x+width]
                        cv2.imwrite(f"D:\\desktop\\Fredrik\\Supplemental_Data_Set_1\\Image_segmentation\\B_TAP_photos_segmentation_in_one\\{file_name}_{chr(i+65)}_{j}.jpg", crop_img)
                        x += move_x

read_path(file_pathname)

# code_#2 color_standard_generate
from PIL import Image
import openpyxl

# 定义起始颜色和结束颜色
start_color = (0, 30, 0 )  # 墨绿色
end_color = (240, 240, 50)   # 黄绿色

# 定义Excel文件名和工作表名
filename = ""
sheetname = "Color standards"

# 定义Excel中起始单元格的位置
start_row = 1
start_col = 1

# 定义颜色图的尺寸和颜色数量
width = 1
height = 300
num_colors = height

# 生成颜色图
gradient = Image.new('RGB', (width, height))
for i in range(num_colors):
    r = start_color[0] + int((i / num_colors) * (end_color[0] - start_color[0]))
    g = start_color[1] + int((i / num_colors) * (end_color[1] - start_color[1]))
    b = start_color[2] + int((i / num_colors) * (end_color[2] - start_color[2]))
    gradient.putpixel((0, i), (r, g, b))

# 将颜色输出到Excel中
wb = openpyxl.Workbook()
ws = wb.active
for i in range(num_colors):
    color = gradient.getpixel((0, i))
    ws.cell(row=start_row+i, column=start_col).value = f"RGB({color[0]}, {color[1]}, {color[2]})"
    ws.cell(row=start_row+i, column=start_col+1).fill = openpyxl.styles.PatternFill(start_color='FF'+"".join([hex(c)[2:].rjust(2, '0').upper() for c in color]), fill_type='solid')
    
# 保存Excel文件
wb.save(filename)
